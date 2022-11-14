import openpyxl
import os


current_folder = os.getcwd()
file = "{}\excel\\book1.xlsx".format(current_folder)

# file -> workbook -> sheets -> rows -> cells
workbook = openpyxl.load_workbook(file)
sheet = workbook["Sheet1"]

row_number = sheet.max_row 
column_number = sheet.max_column

for row in range(1, row_number+1):
  for column in range(1, column_number+1):
    print(sheet.cell(row, column).value, end=" ")
  print()